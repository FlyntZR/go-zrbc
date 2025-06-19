#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Nginx日志分析脚本
统计每个访问链接的平均访问时长
"""

import re
import sys
import os
from collections import defaultdict
from urllib.parse import urlparse, parse_qs
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("警告: 未安装pandas或openpyxl，无法生成Excel文件")
    print("请运行: pip install pandas openpyxl")

def parse_nginx_log_line(line):
    """
    解析nginx日志行
    格式: [时间戳][IP][状态码][响应大小][时间1][时间2]METHOD URL "User-Agent" "Referer"
    """
    # 匹配日志格式的正则表达式 - 处理可能包含换行符的URL
    pattern = r'\[([^\]]+)\]\[([^\]]+)\]\[([^\]]+)\]\[([^\]]+)\]\[([^\]]+)\]\[([^\]]+)\]([A-Z]+)\s+([^"]+)\s+"([^"]*)"\s+"([^"]*)"'
    
    match = re.match(pattern, line.strip())
    if not match:
        return None
    
    timestamp, ip, status_code, response_size, time1, time2, method, url, user_agent, referer = match.groups()
    
    # 清理URL中的换行符和多余空格
    url = url.replace('\n', '').replace('\r', '').strip()
    
    # 提取访问时长（使用第一个时间值）
    try:
        access_time = float(time1)
    except ValueError:
        access_time = 0.0
    
    # 解析URL，提取路径和参数
    try:
        parsed_url = urlparse(url)
        path = parsed_url.path
        query_params = parse_qs(parsed_url.query)
        
        # 构建访问链接标识（路径+主要参数）
        if 'cmd' in query_params:
            cmd = query_params['cmd'][0]
            link_key = f"{path}?cmd={cmd}"
        else:
            link_key = path
    except Exception:
        # 如果URL解析失败，使用原始URL作为key
        link_key = url
    
    return {
        'timestamp': timestamp,
        'ip': ip,
        'status_code': int(status_code),
        'response_size': int(response_size),
        'access_time': access_time,
        'method': method,
        'url': url,
        'link_key': link_key,
        'user_agent': user_agent,
        'referer': referer
    }

def create_excel_report(results, overall_stats, log_file_path, output_file):
    """
    创建Excel报告
    """
    if not EXCEL_AVAILABLE:
        return
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 创建详细统计工作表
    ws_detail = wb.create_sheet("访问链接统计")
    
    # 设置列标题
    headers = ["访问链接", "请求次数", "平均时长(秒)", "最小时长(秒)", "最大时长(秒)", "总时长(秒)"]
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 添加数据
    for row, result in enumerate(results, 2):
        ws_detail.cell(row=row, column=1, value=result['link'])
        ws_detail.cell(row=row, column=2, value=result['count'])
        ws_detail.cell(row=row, column=3, value=result['avg_time'])
        ws_detail.cell(row=row, column=4, value=result['min_time'])
        ws_detail.cell(row=row, column=5, value=result['max_time'])
        ws_detail.cell(row=row, column=6, value=result['total_time'])
    
    # 设置列宽
    ws_detail.column_dimensions['A'].width = 50
    ws_detail.column_dimensions['B'].width = 12
    ws_detail.column_dimensions['C'].width = 15
    ws_detail.column_dimensions['D'].width = 15
    ws_detail.column_dimensions['E'].width = 15
    ws_detail.column_dimensions['F'].width = 15
    
    # 创建总体统计工作表
    ws_summary = wb.create_sheet("总体统计")
    
    # 添加总体统计信息
    summary_data = [
        ["分析时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["日志文件", log_file_path],
        ["总请求数", overall_stats['total_requests']],
        ["解析失败行数", overall_stats['failed_parses']],
        ["唯一访问链接数", overall_stats['unique_links']],
        ["所有请求平均访问时长(秒)", overall_stats['overall_avg']],
        ["最短访问时长(秒)", overall_stats['overall_min']],
        ["最长访问时长(秒)", overall_stats['overall_max']],
    ]
    
    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)
    
    # 设置列宽
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # 保存文件
    wb.save(output_file)
    print(f"Excel报告已保存到: {output_file}")

def analyze_nginx_log(log_file_path):
    """
    分析nginx日志文件
    """
    link_stats = defaultdict(lambda: {'count': 0, 'total_time': 0.0, 'times': []})
    total_requests = 0
    failed_parses = 0
    
    print(f"正在分析日志文件: {log_file_path}")
    print("=" * 80)
    
    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                parsed = parse_nginx_log_line(line)
                if parsed:
                    total_requests += 1
                    link_key = parsed['link_key']
                    access_time = parsed['access_time']
                    
                    link_stats[link_key]['count'] += 1
                    link_stats[link_key]['total_time'] += access_time
                    link_stats[link_key]['times'].append(access_time)
                else:
                    failed_parses += 1
                    
                # 每处理10000行显示进度
                if line_num % 10000 == 0:
                    print(f"已处理 {line_num} 行...")
    
    except FileNotFoundError:
        print(f"错误: 找不到文件 {log_file_path}")
        return
    except Exception as e:
        print(f"读取文件时发生错误: {e}")
        return
    
    # 计算统计结果
    results = []
    for link_key, stats in link_stats.items():
        print(f"link_key: {link_key}, stats: {stats}")
        avg_time = stats['total_time'] / stats['count']
        min_time = min(stats['times'])
        max_time = max(stats['times'])
        
        results.append({
            'link': link_key,
            'count': stats['count'],
            'avg_time': avg_time,
            'min_time': min_time,
            'max_time': max_time,
            'total_time': stats['total_time']
        })
    
    # 按平均访问时长排序
    results.sort(key=lambda x: x['avg_time'], reverse=True)
    
    # 计算总体统计
    overall_stats = {
        'total_requests': total_requests,
        'failed_parses': failed_parses,
        'unique_links': len(results)
    }
    
    if results:
        all_times = []
        for stats in link_stats.values():
            all_times.extend(stats['times'])
        
        overall_stats['overall_avg'] = sum(all_times) / len(all_times)
        overall_stats['overall_min'] = min(all_times)
        overall_stats['overall_max'] = max(all_times)
    else:
        overall_stats['overall_avg'] = 0
        overall_stats['overall_min'] = 0
        overall_stats['overall_max'] = 0
    
    # 输出结果
    print(f"\n分析完成!")
    print(f"总请求数: {total_requests}")
    print(f"解析失败行数: {failed_parses}")
    print(f"唯一访问链接数: {len(results)}")
    print("\n" + "=" * 80)
    print("按平均访问时长排序的统计结果:")
    print("=" * 80)
    
    print(f"{'访问链接':<50} {'请求次数':<10} {'平均时长(秒)':<12} {'最小时长':<10} {'最大时长':<10}")
    print("-" * 100)
    
    for result in results:
        print(f"{result['link']:<50} {result['count']:<10} {result['avg_time']:<12.4f} {result['min_time']:<10.4f} {result['max_time']:<10.4f}")
    
    # 输出总体统计
    print("\n" + "=" * 80)
    print("总体统计:")
    print("=" * 80)
    
    if results:
        print(f"所有请求平均访问时长: {overall_stats['overall_avg']:.4f} 秒")
        print(f"最短访问时长: {overall_stats['overall_min']:.4f} 秒")
        print(f"最长访问时长: {overall_stats['overall_max']:.4f} 秒")
        print(f"总请求数: {total_requests}")
    
    # 生成Excel文件
    if EXCEL_AVAILABLE:
        # 生成输出文件名
        log_filename = os.path.basename(log_file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join("/Users/lld/Downloads", f"nginx_analysis_{timestamp}.xlsx")
        
        create_excel_report(results, overall_stats, log_file_path, output_file)
    else:
        print("\n要生成Excel文件，请安装依赖:")
        print("pip install pandas openpyxl")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python analyze_nginx_log.py <日志文件路径>")
        print("示例: python analyze_nginx_log.py /path/to/nginx.log")
        sys.exit(1)
    
    log_file = sys.argv[1]
    
    # 检查文件是否存在
    if not os.path.exists(log_file):
        print(f"错误: 文件不存在 - {log_file}")
        sys.exit(1)
    
    analyze_nginx_log(log_file) 